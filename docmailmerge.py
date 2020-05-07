from openpyxl import load_workbook #used to import the Excel Data
from datetime import datetime #used to work with date times
#used for merge tags. If there is an error, uninstall and install docx-mailmerge
from mailmerge import MailMerge

# Setting up Excel sheet variables
wb = load_workbook('SampleData.xlsx') #open excel workbook
sheet = wb['SalesOrders'] #Tab to get information
max_row = sheet.max_row #count of all of the rows

# Getting Unique reps. Need to make each of their reports
rep_list = []
for cell_row in range(2 , max_row+1):
    rep = sheet.cell(row = cell_row, column = 3).value
    rep_list.append(rep)
unique_rep_list = list(set(rep_list)) #getting unique list of reps

# For each rep, create their order reports
for rep in unique_rep_list:
    sales_history_list = [] #needed to create the docs dynamically
    raw_subtotal_list = [] #needed to calculate the total

    # Setting up Word document variables. Need to reuse template for each rep
    template_doc = "OrderTemplate.docx"
    word_doc = MailMerge(template_doc)

    for cell_row in range(2 , max_row+1):
        #looping to check the current rep in spreadsheet
        current_rep = sheet.cell(row = cell_row, column = 3).value

        #Checking to see if line item is for rep
        if current_rep == rep:
            #formating date
            #unformatted datetime as a string
            raw_date_time = sheet.cell(row = cell_row, column = 1).value

            #converts datetime back into formatted string
            clean_date_time = raw_date_time.strftime("%m/%d/%Y")

            #Formatting subtotals
            raw_subtotal = sheet.cell(row = cell_row, column = 7).value
            #appending raw number for the total calculation
            raw_subtotal_list.append(raw_subtotal)

            #convert the number into a string and format (example 1,278.25)
            clean_subtotal = "{:,.2f}".format(raw_subtotal)

            #Appending product as a dict into a list, which will
            # be merged as a table
            product_dict = {
            'Date' : clean_date_time,
            'Item' : str(sheet.cell(row = cell_row, column = 4).value),
            'Quantity' : str(sheet.cell(row = cell_row, column = 5).value),
            'Cost' : str(sheet.cell(row = cell_row, column = 6).value),
            'Subtotal' : clean_subtotal
            }

            #Appending dicts to merge as a table
            sales_history_list.append(product_dict)

    # summing raw numbers into a total
    total = sum(raw_subtotal_list)

    # Merging the name and formatting totals
    word_doc.merge(Name = rep, Total = "{:,.2f}".format(total))
    word_doc.merge_rows('Date', sales_history_list) #merge which creates table
    word_doc.write(f'Sales Order for {rep}.docx') #Creates Word doc and names it
